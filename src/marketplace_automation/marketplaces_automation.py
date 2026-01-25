import pandas as pd
from pathlib import Path
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import smtplib
from email.message import EmailMessage

class POReportApp:
    """
    A GUI application to generate Purchase Order (PO) reports for different marketplaces.
    Supports: Blinkit, Flipkart, Swiggy, Zepto (coming soon)
    Includes email functionality to send summary reports.
    """
    
    # Email Configuration
    EMAIL_SENDER = "abhishekwagh420@gmail.com"
    EMAIL_PASSWORD = "bomn ktfx jhct xexy"
    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    DEFAULT_RECIPIENT = "abhishek.wagh@reneecosmetics.in"
    
    def __init__(self, root):
        """Initialize the main window and configure the UI."""
        self.root = root
        self.root.title("PO Report Generator")
        self.root.geometry("460x320")
        self.root.resizable(False, False)
        self.marketplace_var = tk.StringVar(value="Blinkit")
        self.last_summary = {}
        self._build_ui()
    
    def _build_ui(self):
        """Construct the Tkinter widgets for the application."""
        ttk.Label(self.root, text="PO Report Generator", font=("Segoe UI", 16, "bold")).pack(pady=(20, 10))
        ttk.Label(self.root, text="Select Marketplace:").pack(pady=(10, 5))
        self.marketplace_dropdown = ttk.Combobox(
            self.root,
            textvariable=self.marketplace_var,
            values=["Blinkit", "Flipkart", "Swiggy", "Zepto"],
            state="readonly",
            width=25
        )
        self.marketplace_dropdown.pack()
        ttk.Button(
            self.root,
            text="Select CSV/Xlsx and Generate Report",
            command=self.generate_report
        ).pack(pady=(30, 10))
        ttk.Label(
            self.root,
            text="Other marketplaces are coming soon!",
            font=("Segoe UI", 9, "italic"),
            foreground="green"
        ).pack()
        ttk.Label(
            self.root,
            text=f"Last Updated: {datetime.now().strftime('%d-%m-%Y')} | Owner: RENEE-723",
            font=("Segoe UI", 8),
            foreground="gray"
        ).pack(pady=(5, 0))
    
    @staticmethod
    def format_indian(number):
        """Format number in Indian numbering system (lakhs, crores)."""
        s = str(int(number))
        if len(s) <= 3:
            return s
        last3 = s[-3:]
        remaining = s[:-3]
        parts = []
        while len(remaining) > 2:
            parts.insert(0, remaining[-2:])
            remaining = remaining[:-2]
        if remaining:
            parts.insert(0, remaining)
        return ','.join(parts) + ',' + last3
    
    def send_email_summary(self, marketplace, summary_data, tracker_df=None, sku_df=None):
        """
        Send email summary with PO report details.
        
        Args:
            marketplace: Name of the marketplace
            summary_data: Dictionary containing summary statistics
            tracker_df: DataFrame with tracker data (optional)
            sku_df: DataFrame with SKU demand data (optional)
        """
        try:
            # Create email message
            message = EmailMessage()
            message["From"] = self.EMAIL_SENDER
            message["To"] = self.DEFAULT_RECIPIENT
            message["Subject"] = f"📊 Purchase Order Summary: {marketplace} - {datetime.now().strftime('%d-%m-%Y')}"
            
            # Build HTML email body
            html_body = self._build_email_html(marketplace, summary_data, tracker_df, sku_df)
            
            # Set email content
            message.set_content("Please view this email in an HTML-compatible email client.")
            message.add_alternative(html_body, subtype="html")
            
            # Send email
            server = smtplib.SMTP(self.SMTP_SERVER, self.SMTP_PORT)
            server.starttls()
            server.login(self.EMAIL_SENDER, self.EMAIL_PASSWORD)
            server.send_message(message)
            server.quit()
            
            return True
            
        except Exception as e:
            messagebox.showerror("Email Error", f"Failed to send email:\n{str(e)}")
            return False
    
    def _build_email_html(self, marketplace, summary_data, tracker_df, sku_df=None):
        """Build HTML email body with summary, tracker, and SKU data."""
        
        # Extract summary data
        total_pos = summary_data.get('total_pos', 0)
        total_units = summary_data.get('total_units', 0)
        total_value = summary_data.get('total_value', 0)
        min_date = summary_data.get('min_date', 'N/A')
        max_date = summary_data.get('max_date', 'N/A')
        
        # Check if SKU data exists
        has_sku = sku_df is not None and not sku_df.empty
        
        html = f"""
<html>
<head>
<style>
body {{
    font-family: Arial, sans-serif;
    margin: 0;
    padding: 0;
}}
.header {{
    background-color: #4472C4;
    color: white;
    padding: 15px;
    text-align: center;
}}
.summary-table {{
    width: 100%;
    border-collapse: collapse;
    margin: 20px 0;
}}
.summary-table td {{
    padding: 10px;
    border: 1px solid #ddd;
}}
.summary-table td:first-child {{
    background-color: #f2f2f2;
    font-weight: bold;
    width: 200px;
}}
.data-table {{
    width: 100%;
    border-collapse: collapse;
    margin: 20px 0;
}}
.data-table th {{
    background-color: #4472C4;
    color: white;
    padding: 10px;
    text-align: center;
    border: 1px solid #ddd;
}}
.data-table td {{
    padding: 8px;
    text-align: center;
    border: 1px solid #ddd;
}}
.data-table tr:nth-child(even) {{
    background-color: #f9f9f9;
}}
.sku-table {{
    width: 100%;
    border-collapse: collapse;
    margin: 20px 0;
}}
.sku-table th {{
    background-color: #70AD47;
    color: white;
    padding: 10px;
    text-align: center;
    border: 1px solid #ddd;
}}
.sku-table td {{
    padding: 8px;
    text-align: center;
    border: 1px solid #ddd;
}}
.sku-table tr:nth-child(even) {{
    background-color: #f9f9f9;
}}
.footer {{
    text-align: center;
    color: #666;
    font-size: 11px;
    margin-top: 20px;
    padding-top: 10px;
    border-top: 1px solid #ddd;
}}
</style>
</head>
<body>
<div class="header">
<h2 style="margin: 0;">📊 {marketplace} PO Report - {datetime.now().strftime('%d-%m-%Y %H:%M')}</h2>
</div>

<h3 style="margin: 20px 0 10px 0;">Summary Overview</h3>
<table class="summary-table">
<tr>
    <td>Total POs:</td>
    <td>{total_pos}</td>
</tr>
<tr>
    <td>Order Date Range:</td>
    <td>{min_date} to {max_date}</td>
</tr>
<tr>
    <td>Total Units Ordered:</td>
    <td>{self.format_indian(total_units)}</td>
</tr>
<tr>
    <td>Total Order Value:</td>
    <td>₹ {self.format_indian(total_value)}</td>
</tr>
</table>
"""
        
        # Add PO Details table if tracker_df is provided
        if tracker_df is not None and not tracker_df.empty:
            html += '<h3 style="margin: 20px 0 10px 0;">PO Details (All POs)</h3>\n<table class="data-table">\n<tr>'
            
            # Add table headers
            for col in tracker_df.columns:
                html += f"<th>{col}</th>"
            html += "</tr>\n"
            
            # Add table rows (ALL rows - not limited)
            for idx, row in tracker_df.iterrows():
                html += "<tr>"
                for col in tracker_df.columns:
                    cell_value = row[col]
                    
                    # Format dates to DD-MM-YYYY
                    if 'Date' in col and hasattr(cell_value, 'strftime'):
                        cell_value = cell_value.strftime('%d-%m-%Y')
                    # Format PO Value with Indian currency formatting
                    elif col == 'PO Value' or 'total_amount' in col:
                        if isinstance(cell_value, str):
                            clean_value = cell_value.replace('₹', '').replace(',', '').strip()
                            try:
                                numeric_value = float(clean_value)
                                cell_value = f"₹ {self.format_indian(numeric_value)}"
                            except:
                                pass
                        elif isinstance(cell_value, (int, float)):
                            cell_value = f"₹ {self.format_indian(cell_value)}"
                    
                    html += f"<td>{cell_value}</td>"
                html += "</tr>\n"
            
            html += "</table>"
        
        # Add SKU Demand table if sku_df is provided
        if has_sku:
            html += '<h3 style="margin: 20px 0 10px 0;">SKU Demand (All SKUs)</h3>\n<table class="sku-table">\n<tr>'
            
            # Add SKU table headers
            for col in sku_df.columns:
                display_col = str(col).replace('_', ' ').title()
                html += f"<th>{display_col}</th>"
            html += "</tr>\n"
            
            # Add SKU table rows (ALL rows - not limited)
            for idx, row in sku_df.iterrows():
                html += "<tr>"
                for col in sku_df.columns:
                    cell_value = row[col]
                    # Format numbers with Indian formatting if it's the units column
                    if 'unit' in str(col).lower() or 'qty' in str(col).lower():
                        try:
                            cell_value = self.format_indian(int(cell_value))
                        except:
                            pass
                    html += f"<td>{cell_value}</td>"
                html += "</tr>\n"
            
            html += "</table>"
        
        # Add footer
        html += """
<div class="footer">
<p>This is an automated email from PO Report Generator<br>RENEE-723 | © 2026</p>
</div>
</body>
</html>
"""
        
        return html
    
    def show_summary_popup(self, df, marketplace):
        """
        Show summary popup for the given dataframe.
        Works for Blinkit, Flipkart, and Swiggy.
        """
        if marketplace == "Blinkit":
            total_pos = df['po_number'].nunique()
            total_units = df['units_ordered'].sum()
            total_value = df['total_amount'].sum()
            min_date = df['order_date'].min()
            max_date = df['expiry_date'].max()
        elif marketplace in ["Flipkart", "Swiggy"]:
            total_pos = df['PO'].nunique()
            total_units = df['PO Qty'].sum()
            df['PO Value'] = df['PO Value'].replace('[₹,]', '', regex=True).astype(float)
            total_value = df['PO Value'].sum()
            min_date = df['Order Date'].min()
            max_date = df['Expiry Date'].max()
        else:
            total_pos = df.shape[0]
            total_units = 0
            total_value = 0
            min_date = max_date = None
        
        # Format dates to DD-MM-YYYY
        min_date_str = min_date.strftime('%d-%m-%Y') if hasattr(min_date, 'strftime') else str(min_date)
        max_date_str = max_date.strftime('%d-%m-%Y') if hasattr(max_date, 'strftime') else str(max_date)
        
        summary_text = (
            f"📊 SUMMARY REPORT 📊\n\n"
            f"Total POs: {total_pos}\n"
            f"Order Date Range: {min_date_str} to {max_date_str}\n"
            f"Total Units Ordered: {self.format_indian(total_units)}\n"
            f"Total Order Value: ₹ {self.format_indian(total_value)}"
        )
        messagebox.showinfo("PO Summary", summary_text)
        
        # Store summary data for email
        self.last_summary = {
            'total_pos': total_pos,
            'total_units': total_units,
            'total_value': total_value,
            'min_date': min_date_str,
            'max_date': max_date_str
        }
    
    def generate_report(self):
        """Main function to generate PO reports for different marketplaces."""
        marketplace = self.marketplace_var.get()
        
        # Ask user to select CSV/Excel file
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return
        
        try:
            # --- Blinkit Logic --- 
            if marketplace == "Blinkit":
                df = pd.read_csv(file_path)
                df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
                df['po_number'] = df['po_number'].astype(str).str.replace(r'\.0$', '', regex=True)
                df['order_date'] = pd.to_datetime(df['order_date'], dayfirst=True, errors='coerce').dt.date
                df['expiry_date'] = pd.to_datetime(df['expiry_date'], dayfirst=True, errors='coerce').dt.date
                
                tracker_summary = df.groupby(['po_number', 'facility_name'], as_index=False).agg({
                    'order_date': 'first',
                    'expiry_date': 'first',
                    'total_amount': 'sum',
                    'units_ordered': 'sum'
                })
                tracker_summary.insert(0, 'marketplace', marketplace)
                tracker_summary['total_amount'] = tracker_summary['total_amount'].apply(
                    lambda x: f"₹ {self.format_indian(x)}"
                )
                tracker_summary = tracker_summary.sort_values('facility_name', ascending=True)
                
                df['upc'] = df['upc'].apply(lambda x: str(int(float(x))))
                sku_summary = df.groupby(['upc', 'name'], as_index=False).agg({'units_ordered': 'sum'})
                sku_summary.rename(columns={'units_ordered': 'total_units'}, inplace=True)
                sku_summary = sku_summary.sort_values(by='total_units', ascending=False)
                
                timestamp = datetime.now().strftime("%d_%m_%Y__%H_%M_%S")
                output_file = Path(file_path).parent / f"{marketplace}_PO_Report_{timestamp}.xlsx"
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    tracker_summary.to_excel(writer, sheet_name='PO Tracker', index=False)
                    sku_summary.to_excel(writer, sheet_name='SKU Summary', index=False)
                    
                    from openpyxl.styles import Alignment
                    from openpyxl.utils import get_column_letter
                    workbook = writer.book
                    
                    for sheet_name in ['PO Tracker', 'SKU Summary']:
                        ws = workbook[sheet_name]
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        for col in ws.columns:
                            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                            col_letter = get_column_letter(col[0].column)
                            ws.column_dimensions[col_letter].width = max_length + 2
                
                self.show_summary_popup(df, marketplace)
                messagebox.showinfo("Success", f"Report created successfully at:\n{output_file}")
                
                # Ask if user wants to send email BEFORE opening file
                if messagebox.askyesno("Send Email", "Do you want to send this summary via email?"):
                    if self.send_email_summary(marketplace, self.last_summary, tracker_summary, sku_summary):
                        messagebox.showinfo("Email Sent", f"Summary sent successfully to {self.DEFAULT_RECIPIENT}")
                
                if messagebox.askyesno("Open File", "Do you want to open the generated Excel file?"):
                    os.startfile(output_file)
                    self.root.destroy()
            
            # --- Flipkart Logic ---
            elif marketplace == "Flipkart":
                if str(file_path).endswith('.csv'):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)
                
                cols_map = {
                    "Purchase Order ID": "PO",
                    "Origin Warehouse": "Location",
                    "Order Date": "Order Date",
                    "Expiry Date": "Expiry Date",
                    "Total Amount": "PO Value",
                    "Total Ordered Quantity": "PO Qty"
                }
                df = df[list(cols_map.keys())].rename(columns=cols_map)
                
                df["Order Date"] = pd.to_datetime(df["Order Date"], dayfirst=True, errors="coerce").dt.date
                df["Expiry Date"] = pd.to_datetime(df["Expiry Date"], dayfirst=True, errors="coerce").dt.date
                
                FLIPKART_ALPHA_LOCS = [
                    "ban_ven_wh_nl_01nl",
                    "frk_bts",
                    "gur_san_wh_nl_01nl",
                    "malur_bts",
                    "nad_har_wh_kl_nl_01nl"
                ]
                
                df["Marketplace"] = df["Location"].apply(
                    lambda x: "Flipkart Alpha" if x in FLIPKART_ALPHA_LOCS else "Flipkart Hyperlocal"
                )
                
                tracker_summary = df[["Marketplace", "PO", "Location", "Order Date", "Expiry Date", "PO Value", "PO Qty"]].copy()
                tracker_summary = tracker_summary.sort_values("Location", ascending=True)
                
                timestamp = datetime.now().strftime("%d_%m_%Y__%H_%M_%S")
                output_file = Path(file_path).parent / f"{marketplace}_PO_Report_{timestamp}.xlsx"
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    tracker_summary.to_excel(writer, sheet_name='PO Tracker', index=False)
                    
                    from openpyxl.styles import Alignment
                    from openpyxl.utils import get_column_letter
                    workbook = writer.book
                    
                    ws = workbook['PO Tracker']
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    for col in ws.columns:
                        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = max_length + 2
                
                self.show_summary_popup(tracker_summary, marketplace)
                messagebox.showinfo("Success", f"Report created successfully at:\n{output_file}")
                
                # Ask if user wants to send email BEFORE opening file (No SKU for Flipkart)
                if messagebox.askyesno("Send Email", "Do you want to send this summary via email?"):
                    if self.send_email_summary(marketplace, self.last_summary, tracker_summary, None):
                        messagebox.showinfo("Email Sent", f"Summary sent successfully to {self.DEFAULT_RECIPIENT}")
                
                if messagebox.askyesno("Open File", "Do you want to open the generated Excel file?"):
                    os.startfile(output_file)
                    self.root.destroy()
            
            # --- Swiggy Logic ---
            elif marketplace == "Swiggy":
                if str(file_path).endswith('.csv'):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)
                
                df.columns = df.columns.str.strip().str.replace(" ", "").str.upper()
                df = df[df['STATUS'] == 'CONFIRMED']
                
                df['POCREATEDAT'] = pd.to_datetime(df['POCREATEDAT'], dayfirst=True, errors='coerce').dt.date
                df['POEXPIRYDATE'] = pd.to_datetime(df['POEXPIRYDATE'], dayfirst=True, errors='coerce').dt.date
                
                # Aggregate PO Tracker
                tracker_summary = df.groupby(['PONUMBER', 'CITY'], as_index=False).agg({
                    'POCREATEDAT': 'first',
                    'POEXPIRYDATE': 'first',
                    'POLINEVALUEWITHTAX': 'sum',
                    'ORDEREDQTY': 'sum'
                })
                tracker_summary.insert(0, 'Marketplace', 'Swiggy')
                tracker_summary.rename(columns={
                    'PONUMBER': 'PO',
                    'CITY': 'Location',
                    'POCREATEDAT': 'Order Date',
                    'POEXPIRYDATE': 'Expiry Date',
                    'POLINEVALUEWITHTAX': 'PO Value',
                    'ORDEREDQTY': 'PO Qty'
                }, inplace=True)
                
                # Create SKU Summary (similar to Blinkit)
                # Fix EAN conversion - handle scientific notation properly
                def safe_ean_convert(x):
                    if pd.isna(x):
                        return ""
                    try:
                        # Convert to string first, then to float, then to int
                        return str(int(float(str(x))))
                    except:
                        return str(x)
                
                df['EAN'] = df['EAN'].apply(safe_ean_convert)
                
                sku_summary = df.groupby(['EAN', 'SKUDESCRIPTION'], as_index=False).agg({'ORDEREDQTY': 'sum'})
                sku_summary.rename(columns={
                    'EAN': 'upc',
                    'SKUDESCRIPTION': 'name',
                    'ORDEREDQTY': 'total_units'
                }, inplace=True)
                sku_summary = sku_summary.sort_values(by='total_units', ascending=False)
                
                # Format PO Value for tracker
                tracker_summary['PO Value'] = tracker_summary['PO Value'].apply(lambda x: f"₹ {self.format_indian(x)}")
                
                timestamp = datetime.now().strftime("%d_%m_%Y__%H_%M_%S")
                
                # Create main Swiggy report (like Blinkit)
                main_output_file = Path(file_path).parent / f"{marketplace}_PO_Report_{timestamp}.xlsx"
                
                with pd.ExcelWriter(main_output_file, engine='openpyxl') as writer:
                    tracker_summary.to_excel(writer, sheet_name='PO Tracker', index=False)
                    sku_summary.to_excel(writer, sheet_name='SKU Summary', index=False)
                    
                    from openpyxl.styles import Alignment
                    from openpyxl.utils import get_column_letter
                    workbook = writer.book
                    
                    for sheet_name in ['PO Tracker', 'SKU Summary']:
                        ws = workbook[sheet_name]
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        for col in ws.columns:
                            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                            col_letter = get_column_letter(col[0].column)
                            ws.column_dimensions[col_letter].width = max_length + 2
                
                # Create separate folder for individual PO workbooks
                output_folder = Path(file_path).parent / f"Swiggy_PO_Workbooks_{timestamp}"
                output_folder.mkdir(parents=True, exist_ok=True)
                
                def format_sheet(ws, ean_columns=None):
                    from openpyxl.styles import Alignment
                    from openpyxl.utils import get_column_letter
                    
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    if ean_columns:
                        for col_letter in ean_columns:
                            for cell in ws[col_letter]:
                                cell.number_format = '@'
                                if cell.value and str(cell.value).replace('.', '').isdigit():
                                    cell.value = str(int(float(cell.value)))
                    
                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 4, 50)
                        ws.column_dimensions[col_letter].width = adjusted_width
                
                from openpyxl import Workbook
                from openpyxl.styles import Font
                
                for po_number in tracker_summary['PO']:
                    po_data = df[df['PONUMBER'] == po_number]
                    city = po_data['CITY'].iloc[0]
                    wb = Workbook()
                    
                    ws_po = wb.active
                    ws_po.title = "PO Sheet"
                    ws_po.append(['EAN', 'SKUDESCRIPTION', 'UNITBASEDCOST', 'ORDEREDQTY'])
                    for idx, row in po_data.iterrows():
                        # Safe EAN conversion
                        try:
                            ean_value = str(int(float(row['EAN']))) if pd.notna(row['EAN']) else ""
                        except:
                            ean_value = str(row['EAN']) if pd.notna(row['EAN']) else ""
                        
                        ws_po.append([
                            ean_value,
                            row['SKUDESCRIPTION'],
                            row['UNITBASEDCOST'],
                            row['ORDEREDQTY']
                        ])
                    format_sheet(ws_po, ean_columns=['A'])
                    
                    ws_scan = wb.create_sheet("Scan Sheet")
                    ws_scan.append(["Scan", "EAN", "Title", "PO Qty", "Qty", "Box"])
                    for r in range(2, len(po_data) + 20):
                        ws_scan[f'B{r}'] = f'=IF(A{r}="","",TEXT(INDEX(\'PO Sheet\'!A:A,MATCH(A{r},\'PO Sheet\'!A:A,0)),"0"))'
                        ws_scan[f'C{r}'] = f'=IF(A{r}="","",INDEX(\'PO Sheet\'!B:B,MATCH(A{r},\'PO Sheet\'!A:A,0)))'
                        ws_scan[f'D{r}'] = f'=IF(A{r}="","",INDEX(\'PO Sheet\'!D:D,MATCH(A{r},\'PO Sheet\'!A:A,0)))'
                    format_sheet(ws_scan, ean_columns=['A', 'B'])
                    
                    ws_packing = wb.create_sheet("Packing Slip")
                    ws_packing['A1'] = f"{po_number} Swiggy {city}"
                    ws_packing['A1'].font = Font(bold=True, size=12)
                    ws_packing.merge_cells('A1:D1')
                    ws_packing.append(["Scan", "Title", "Box", "Total"])
                    for r in range(3, len(po_data) + 3):
                        scan_row = r - 1
                        ws_packing[f'A{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",TEXT(\'Scan Sheet\'!A{scan_row},"0"))'
                        ws_packing[f'B{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",\'Scan Sheet\'!C{scan_row})'
                        ws_packing[f'C{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",\'Scan Sheet\'!F{scan_row})'
                        ws_packing[f'D{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",\'Scan Sheet\'!D{scan_row})'
                    total_row = len(po_data) + 3
                    ws_packing[f'D{total_row}'] = f"=SUBTOTAL(109,D3:D{total_row-1})"
                    ws_packing[f'D{total_row}'].font = Font(bold=True)
                    format_sheet(ws_packing, ean_columns=['A'])
                    
                    wb_path = output_folder / f"{po_number}.xlsx"
                    wb.save(wb_path)
                
                self.show_summary_popup(tracker_summary, marketplace)
                
                messagebox.showinfo(
                    "Success", 
                    f"Main report created: {main_output_file.name}\n\n"
                    f"Individual PO workbooks created in:\n{output_folder.name}"
                )
                
                # Ask if user wants to send email BEFORE opening files
                if messagebox.askyesno("Send Email", "Do you want to send this summary via email?"):
                    if self.send_email_summary(marketplace, self.last_summary, tracker_summary, sku_summary):
                        messagebox.showinfo("Email Sent", f"Summary sent successfully to {self.DEFAULT_RECIPIENT}")
                
                if messagebox.askyesno("Open File", "Do you want to open the main Excel report?"):
                    os.startfile(main_output_file)
                
                if messagebox.askyesno("Open Folder", "Do you want to open the PO workbooks folder?"):
                    os.startfile(output_folder)
            
            # --- Zepto (Future Placeholder) ---
            elif marketplace == "Zepto":
                messagebox.showinfo("Info", "Zepto integration is coming soon!")
        
        except Exception as e:
            messagebox.showerror("Error", f"Something went wrong:\n{e}")

def main():
    root = tk.Tk()
    app = POReportApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()