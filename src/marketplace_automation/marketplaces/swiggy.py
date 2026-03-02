import pandas as pd
from pathlib import Path
from datetime import datetime
import os
from utils import format_indian, safe_ean_convert

def process_swiggy(file_path):
    marketplace = "Swiggy"
    if str(file_path).endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)

    df.columns = df.columns.str.strip().str.replace(" ", "").str.upper()
    df = df[df['STATUS'] == 'CONFIRMED']

    df['POCREATEDAT'] = pd.to_datetime(df['POCREATEDAT'], dayfirst=True, errors='coerce').dt.strftime('%d-%m-%Y')
    df['POEXPIRYDATE'] = pd.to_datetime(df['POEXPIRYDATE'], dayfirst=True, errors='coerce').dt.strftime('%d-%m-%Y')

    # Aggregate PO Tracker
    tracker_summary = df.groupby(['PONUMBER', 'FACILITYNAME'], as_index=False).agg({
        'POCREATEDAT': 'first',
        'POEXPIRYDATE': 'first',
        'POLINEVALUEWITHTAX': 'sum',
        'ORDEREDQTY': 'sum'
    })
    tracker_summary.insert(0, 'Marketplace', 'Swiggy')
    tracker_summary.rename(columns={
        'PONUMBER': 'PO',
        'FACILITYNAME': 'Location',
        'POCREATEDAT': 'Order Date',
        'POEXPIRYDATE': 'Expiry Date',
        'POLINEVALUEWITHTAX': 'PO Value',
        'ORDEREDQTY': 'PO Qty'
    }, inplace=True)

    # Create SKU Summary (similar to Blinkit)
    df['EAN'] = df['EAN'].apply(safe_ean_convert)

    sku_summary = df.groupby(['EAN', 'SKUDESCRIPTION'], as_index=False).agg({'ORDEREDQTY': 'sum'})
    sku_summary.rename(columns={
        'EAN': 'upc',
        'SKUDESCRIPTION': 'name',
        'ORDEREDQTY': 'total_units'
    }, inplace=True)
    sku_summary = sku_summary.sort_values(by='total_units', ascending=False)

    timestamp = datetime.now().strftime("%d_%m_%Y__%H_%M_%S")

    # Create main Swiggy report (like Blinkit)
    main_output_file = Path(file_path).parent / f"{marketplace}_PO_Report_{timestamp}.xlsx"

    # Create a copy for Excel export with formatted amounts
    tracker_summary_excel = tracker_summary.copy()
    tracker_summary_excel['PO Value'] = tracker_summary_excel['PO Value'].apply(lambda x: f"₹ {format_indian(x)}")

    with pd.ExcelWriter(main_output_file, engine='openpyxl') as writer:
        tracker_summary_excel.to_excel(writer, sheet_name='PO Tracker', index=False)
        sku_summary.to_excel(writer, sheet_name='SKU Summary', index=False)

        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
        workbook = writer.book

        for sheet_name in ['PO Tracker', 'SKU Summary']:
            ws = workbook[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_length + 2

    # Create separate folder for individual PO workbooks
    output_folder = Path(file_path).parent / f"Swiggy_PO_Workbooks_{timestamp}"
    output_folder.mkdir(parents=True, exist_ok=True)

    def format_sheet(ws, ean_columns=None):
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        if ean_columns:
            for col_letter in ean_columns:
                for cell in ws[col_letter]:
                    cell.number_format = '@'
                    if cell.value and str(cell.value).replace('.', '').isdigit():
                        cell.value = str(int(float(cell.value)))

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    from openpyxl import Workbook
    from openpyxl.styles import Font

    for po_number in tracker_summary['PO']:
        po_data = df[df['PONUMBER'] == po_number]
        # city = po_data['CITY'].iloc[0]
        facility = po_data['FACILITYNAME'].iloc[0]
        wb = Workbook()

        ws_po = wb.active
        ws_po.title = "PO Sheet"
        ws_po.append(['EAN', 'SKUDESCRIPTION', 'UNITBASEDCOST', 'ORDEREDQTY'])
        for idx, row in po_data.iterrows():
            # Safe EAN conversion
            try:
                ean_value = str(int(float(row['EAN']))) if pd.notna(row['EAN']) else ""
            except:
                ean_value = str(row['EAN']) if pd.notna(row['EAN']) else ""

            ws_po.append([
                ean_value,
                row['SKUDESCRIPTION'],
                row['UNITBASEDCOST'],
                row['ORDEREDQTY']
            ])
        format_sheet(ws_po, ean_columns=['A'])

        ws_scan = wb.create_sheet("Scan Sheet")
        ws_scan.append(["Scan", "EAN", "Title", "PO Qty", "Qty", "Box"])
        for r in range(2, len(po_data) + 20):
            ws_scan[f'B{r}'] = f'=IF(A{r}="","",TEXT(INDEX(\'PO Sheet\'!A:A,MATCH(A{r},\'PO Sheet\'!A:A,0)),"0"))'
            ws_scan[f'C{r}'] = f'=IF(A{r}="","",INDEX(\'PO Sheet\'!B:B,MATCH(A{r},\'PO Sheet\'!A:A,0)))'
            ws_scan[f'D{r}'] = f'=IF(A{r}="","",INDEX(\'PO Sheet\'!D:D,MATCH(A{r},\'PO Sheet\'!A:A,0)))'
        format_sheet(ws_scan, ean_columns=['A', 'B'])

        ws_packing = wb.create_sheet("Packing Slip")
        # ws_packing['A1'] = f"{po_number} Swiggy {city}"
        ws_packing['A1'] = f"{po_number} Swiggy {facility}"
        ws_packing['A1'].font = Font(bold=True, size=12)
        ws_packing.merge_cells('A1:D1')
        ws_packing.append(["Scan", "Title", "Box", "Total"])
        for r in range(3, len(po_data) + 3):
            scan_row = r - 1
            ws_packing[f'A{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",TEXT(\'Scan Sheet\'!A{scan_row},"0"))'
            ws_packing[f'B{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",\'Scan Sheet\'!C{scan_row})'
            ws_packing[f'C{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",\'Scan Sheet\'!F{scan_row})'
            ws_packing[f'D{r}'] = f'=IF(\'Scan Sheet\'!A{scan_row}="","",\'Scan Sheet\'!D{scan_row})'
        total_row = len(po_data) + 3
        ws_packing[f'D{total_row}'] = f"=SUBTOTAL(109,D3:D{total_row-1})"
        ws_packing[f'D{total_row}'].font = Font(bold=True)
        format_sheet(ws_packing, ean_columns=['A'])

        wb_path = output_folder / f"{po_number}.xlsx"
        wb.save(wb_path)

    return {
        "marketplace": marketplace,
        "df": df,
        "tracker_df": tracker_summary,
        "sku_df": sku_summary,
        "output_file": main_output_file,
        "output_folder": output_folder,
        "has_sku_data": True,
        "sku_count": len(sku_summary)
    }
