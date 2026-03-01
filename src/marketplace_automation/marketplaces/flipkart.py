import pandas as pd
from pathlib import Path
from datetime import datetime
import os
from tkinter import messagebox

def process_flipkart(app, file_path):
    marketplace = "Flipkart"
    if str(file_path).endswith('.csv'):
        df = pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path)

    cols_map = {
        "Purchase Order ID": "PO",
        "Origin Warehouse": "Location",
        "Order Date": "Order Date",
        "Expiry Date": "Expiry Date",
        "Total Amount": "PO Value",
        "Total Ordered Quantity": "PO Qty"
    }
    df = df[list(cols_map.keys())].rename(columns=cols_map)

    df["Order Date"] = pd.to_datetime(df["Order Date"], errors="coerce").dt.strftime('%d-%m-%Y')
    df["Expiry Date"] = pd.to_datetime(df["Expiry Date"], errors="coerce").dt.strftime('%d-%m-%Y')

    FLIPKART_ALPHA_LOCS = [
        "bhi_pad_wh_nl_01nl",
        "ban_ven_wh_nl_01nl",
        "frk_bts",
        "gur_san_wh_nl_01nl",
        "malur_bts",
        "nad_har_wh_kl_nl_01nl"
    ]

    df["Marketplace"] = df["Location"].apply(
        lambda x: "Flipkart Alpha" if x in FLIPKART_ALPHA_LOCS else "Flipkart Hyperlocal"
    )

    tracker_summary = df[["Marketplace", "PO", "Location", "Order Date", "Expiry Date", "PO Value", "PO Qty"]].copy()
    tracker_summary = tracker_summary.sort_values("Location", ascending=True)

    timestamp = datetime.now().strftime("%d_%m_%Y__%H_%M_%S")
    output_file = Path(file_path).parent / f"{marketplace}_PO_Report_{timestamp}.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        tracker_summary.to_excel(writer, sheet_name='PO Tracker', index=False)

        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
        workbook = writer.book

        ws = workbook['PO Tracker']
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

    app.show_summary_popup(tracker_summary, marketplace, has_sku_data=False, sku_count=0)
    messagebox.showinfo("Success", f"Report created successfully at:\n{output_file}")

    # Ask if user wants to send email BEFORE opening file (No SKU for Flipkart)
    if messagebox.askyesno("Send Email", "Do you want to send this summary via email?"):
        if app.send_email_summary(marketplace, app.last_summary, tracker_summary, None):
            # Build recipient info for success message
            recipient_info = f"To: {app.DEFAULT_RECIPIENT}"
            if app.CC_RECIPIENTS:
                cc_list = "\n".join([f"  • {email}" for email in app.CC_RECIPIENTS])
                recipient_info += f"\n\nCC:\n{cc_list}"

            messagebox.showinfo("Email Sent", f"Summary sent successfully to:\n\n{recipient_info}")

    if messagebox.askyesno("Open File", "Do you want to open the generated Excel file?"):
        os.startfile(output_file)
        app.root.destroy()
