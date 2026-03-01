import pandas as pd
from pathlib import Path
from datetime import datetime
import os
from utils import format_indian

def process_blinkit(file_path):
    marketplace = "Blinkit"
    df = pd.read_csv(file_path)
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    df['po_number'] = df['po_number'].astype(str).str.replace(r'\.0$', '', regex=True)
    df['order_date'] = pd.to_datetime(df['order_date'], dayfirst=True, errors='coerce')
    df['expiry_date'] = pd.to_datetime(df['expiry_date'], dayfirst=False, errors='coerce')

    df['order_date'] = df['order_date'].dt.strftime('%d-%m-%Y')
    df['expiry_date'] = df['expiry_date'].dt.strftime('%d-%m-%Y')
    tracker_summary = df.groupby(['po_number', 'facility_name'], as_index=False).agg({
        'order_date': 'first',
        'expiry_date': 'first',
        'total_amount': 'sum',
        'units_ordered': 'sum'
    })
    tracker_summary.insert(0, 'marketplace', marketplace)
    tracker_summary['total_amount'] = tracker_summary['total_amount'].apply(
        lambda x: f"₹ {format_indian(x)}"
    )
    tracker_summary = tracker_summary.sort_values('facility_name', ascending=True)

    df['upc'] = df['upc'].apply(lambda x: str(int(float(x))))
    sku_summary = df.groupby(['upc', 'name'], as_index=False).agg({'units_ordered': 'sum'})
    sku_summary.rename(columns={'units_ordered': 'total_units'}, inplace=True)
    sku_summary = sku_summary.sort_values(by='total_units', ascending=False)

    timestamp = datetime.now().strftime("%d_%m_%Y__%H_%M_%S")
    output_file = Path(file_path).parent / f"{marketplace}_PO_Report_{timestamp}.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        tracker_summary.to_excel(writer, sheet_name='PO Tracker', index=False)
        sku_summary.to_excel(writer, sheet_name='SKU Summary', index=False)

        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
        workbook = writer.book

        for sheet_name in ['PO Tracker', 'SKU Summary']:
            ws = workbook[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_length + 2

    return {
        "marketplace": marketplace,
        "df": df,
        "tracker_df": tracker_summary,
        "sku_df": sku_summary,
        "output_file": output_file,
        "has_sku_data": True,
        "sku_count": len(sku_summary)
    }
